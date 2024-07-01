# -*- coding: utf-8 -*-
"""
文 件 名: openpyxl_example2.py
文件描述: 检查xlxs表格的重复数据
作    者: HanKin
创建日期: 2023.09.21
修改日期：2023.09.21

Copyright (c) 2023 HanKin. All rights reserved.
"""

import os
import time
import openpyxl
from collections import Counter

device_dir = "./"

def test():
    """
    检查xlxs表格的重复数据
    :param
    :return
    """
    COMPATIBILITY_LIST_FILENAME = "compatibility_list.xlsx"
    
    # 打开Excel文件
    workbook = openpyxl.load_workbook(os.path.join(device_dir, COMPATIBILITY_LIST_FILENAME))
    
    # 获取工作表
    sheet = workbook.active

    # 获取工作表的大小
    print("{} excel file table shape({}, {})".format(COMPATIBILITY_LIST_FILENAME, sheet.max_row, sheet.max_column))

    # 遍历第一行找到vid_pid列号
    vid_pid_col_idx = 0
    for i in range(1, sheet.max_column+1):
        if sheet.cell(1, i).value == "vid_pid":
            vid_pid_col_idx = i
            break
    if vid_pid_col_idx == 0:
        print("{} excel file is not found vid_pid column".format(COMPATIBILITY_LIST_FILENAME))
        return 

    all_vid_pids = []
    for i in range(2, sheet.max_row+1):
        vid_pids = sheet.cell(i, vid_pid_col_idx).value.split(';')
        
        # 筛选内容过多的行
        suggestion = sheet.cell(i, vid_pid_col_idx + 1).value
        if len(suggestion) > 200:
            print("index {}".format(i))
        
        for vid_pid in vid_pids:
            all_vid_pids.append(vid_pid.lower())
    print("all_vid_pids size {}".format(len(all_vid_pids)))
    
    my_counter = Counter(all_vid_pids)
    duplicates = [item for item, count in my_counter.items() if count > 1]
    print(duplicates)
    
    for item in duplicates:
        for i in range(2, sheet.max_row+1):
            vid_pids = sheet.cell(i, vid_pid_col_idx).value.split(';')
            for vid_pid in vid_pids:
                if item.lower() == vid_pid.lower():
                    print("found {} usb device, index {} model {} vendor {} usb_method {}".format(item,
                             i, sheet.cell(i, 1).value, sheet.cell(i, 2).value, sheet.cell(i, 4).value))
                    break
        print("")
    
    # 关闭Excel文件
    workbook.close()

def practice(device_id):
    """
    检查xlxs表格的重复数据
    :param
    :return
    """
    COMPATIBILITY_LIST_FILENAME = "compatibility_list.xlsx"
    
    # 打开Excel文件
    workbook = openpyxl.load_workbook(os.path.join(device_dir, COMPATIBILITY_LIST_FILENAME))
    
    # 获取工作表
    sheet = workbook.active

    # 获取工作表的大小
    print("{} excel file table shape({}, {})".format(COMPATIBILITY_LIST_FILENAME, sheet.max_row, sheet.max_column))

    # 遍历第一行找到vid_pid列号
    vid_pid_col_idx = 0
    for i in range(1, sheet.max_column+1):
        if sheet.cell(1, i).value == "vid_pid":
            vid_pid_col_idx = i
            break
    if vid_pid_col_idx == 0:
        print("{} excel file is not found vid_pid column".format(COMPATIBILITY_LIST_FILENAME))
        return 

    device_id_idx = []
    suggestion = ""
    for i in range(2, sheet.max_row+1):
        vid_pids = sheet.cell(i, vid_pid_col_idx).value.split(';')
        is_found = False
        for vid_pid in vid_pids:
            if device_id.lower() == vid_pid.lower():
                print("found {} usb device, index {} model {} vendor {} usb_method {}".format(device_id,
                         i, sheet.cell(i, 1).value, sheet.cell(i, 2).value, sheet.cell(i, 4).value))
                device_id_idx.append(sheet.cell(i, 1).value)
                if not is_found:
                    suggestion = sheet.cell(i, vid_pid_col_idx+1).value
                    is_found = True
    
    print(len(device_id_idx))
    print(suggestion)
    
    # 关闭Excel文件
    workbook.close()

def main():
    """主函数
    """

    #test()
    practice("2b16:bcd2")

if __name__ == '__main__':
    """程序入口
    """
    
    #os.system('chcp 936 & cls')
    print('******** starting ********')
    start_time = time.time()
    
    main()

    end_time = time.time()
    print('process spend {} s.\n'.format(round(end_time - start_time, 3)))
