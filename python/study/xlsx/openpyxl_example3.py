# -*- coding: utf-8 -*-
"""
文 件 名: openpyxl_example3.py
文件描述: 定位python脚本内存泄露问题
备    注：需要使用close关闭句柄
作    者: HanKin
创建日期: 2024.06.14
修改日期：2024.06.14

Copyright (c) 2024 HanKin. All rights reserved.
"""

import os
import time
import openpyxl
    
def test():
    """
    """
    
    source_path = 'test.xlsx'
    
    # 打开Excel文件
    if not os.path.exists(source_path):
        workbook = openpyxl.Workbook(source_path)
        sheet = workbook.create_sheet('Sheet1')
        
    else:
        workbook = openpyxl.load_workbook(source_path)
        # 获取工作表
        #sheet = workbook.active
        sheet = workbook['Sheet1'] # 获取 Sheet1 对象
        
        # 获取工作表的大小
        print(sheet.max_row, sheet.max_column)
    
    # 新建的能这样添加不能以下面的方式
    lis = ['张三','在家','地址','学号','手机','昵称']
    sheet.append(lis)
 
    # 读/写  B2 单元格数据
    sheet['B2'].value = 12
    cell_value = sheet['B2'].value
    print('B2 ', cell_value)

    # 读/写 2行 3列 数据
    sheet.cell(2, 3).value = 34
    cell_value = sheet.cell(2, 3).value
    print('2行3列 ', cell_value)
    
    # 保存
    workbook.save(source_path)
    
    # 关闭Excel文件
    #workbook.close()
    
def memory_leak():
    """鉴定完毕，没有内存泄露情况
    即使你没有显式地调用 workbook.close() 方法关闭工作簿，openpyxl 库通常会在工作簿对象被垃圾回收时自动关闭工作簿。这意味着在大多数情况下，你不需要手动关闭工作簿来避免内存泄漏。
    """
    source_path = 'test.xlsx'
    
    # 打开Excel文件
    if os.path.exists(source_path):
        workbook = openpyxl.load_workbook(source_path)
        # 获取工作表
        #sheet = workbook.active
        sheet = workbook['Sheet1'] # 获取 Sheet1 对象
        
        # 获取工作表的大小
        #print(sheet.max_row, sheet.max_column)
    
        cell_value = sheet.cell(2, 3).value
        print('{} 2行3列: {}'.format(time.strftime("%Y-%m-%d %H:%M:%S"), cell_value))
        
        # 保存
        workbook.save(source_path)
        
        # 关闭Excel文件
        #workbook.close()
    
def main():
    """主函数
    """

    #test()
    while True:
        memory_leak()
        #time.sleep(0.5)

if __name__ == '__main__':
    """程序入口
    """
    
    #os.system('chcp 936 & cls')
    print('******** starting ********')
    start_time = time.time()
    
    main()

    end_time = time.time()
    print('process spend {} s.\n'.format(round(end_time - start_time, 3)))
