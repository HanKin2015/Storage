# -*- coding: utf-8 -*-
"""
文 件 名: openpyxl_example.py
文件描述: 使用openpyxl打开xlxs表格
备    注：需要使用close关闭句柄
作    者: HanKin
创建日期: 2023.09.05
修改日期：2023.09.14

Copyright (c) 2023 HanKin. All rights reserved.
"""

import os
import time
import openpyxl

def traverse():
    """遍历列的时候会卡顿一下
    """
    
    # 打开Excel文件
    workbook = openpyxl.load_workbook('外设助手改进/外设助手改进专项.xlsx')

    # 获取工作表
    #sheet = workbook.active
    sheet = workbook['兼容性列表']
    
    # 获取工作表的大小（不准确，输出1048554 10）
    print(sheet.max_row, sheet.max_column)

    # 读取单元格数据
    cell_value = sheet['A1'].value

    # 遍历行
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=5)):
        print('************ {} ************'.format(index))
        for cell in row:
            if cell.value is None:
                break
            print(cell.value)
        print('')
    print('遍历行完成')

    # 遍历列
    for index, column in enumerate(sheet.iter_cols(min_col=1, max_col=5)):
        print('************ {} ************'.format(index))
        for cell in column:
            if cell.value is None:
                break
            print(cell.value)
        print('')
    print('遍历列完成')
    
    # 关闭Excel文件
    workbook.close()
    
def test():
    """更换方式遍历
    """
    
    # 打开Excel文件
    workbook = openpyxl.load_workbook('外设助手改进/外设助手改进专项.xlsx')
    
    # 获取工作表
    #sheet = workbook.active
    sheet = workbook['兼容性列表']
    
    # 获取工作表的大小（不准确，输出1048554 10）
    print(sheet.max_row, sheet.max_column)
    
    # 遍历列
    vid_pid = []
    for i in sheet['E']:
        vid_pid.append(i.value)
        #print(i.value)
    print(len(vid_pid))
    
    # 关闭Excel文件
    workbook.close()
    
def main():
    """主函数
    """
    
    traverse()
    test()

if __name__ == '__main__':
    """程序入口
    """
    
    #os.system('chcp 936 & cls')
    print('******** starting ********')
    start_time = time.time()
    
    main()

    end_time = time.time()
    print('process spend {} s.\n'.format(round(end_time - start_time, 3)))
