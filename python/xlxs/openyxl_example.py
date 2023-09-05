# -*- coding: utf-8 -*-
"""
文 件 名: openyxl_example.py
文件描述: 使用openyxl打开xlxs表格
作    者: HanKin
创建日期: 2023.09.05
修改日期：2023.09.05

Copyright (c) 2023 HanKin. All rights reserved.
"""

import os
import time
import openpyxl

def main():
    """主函数
    """
    
    # 打开Excel文件
    workbook = openpyxl.load_workbook('外设助手改进/外设助手改进专项.xlsx')

    # 获取工作表
    #sheet = workbook.active
    sheet = workbook['兼容性列表']
    
    # 获取工作表的大小
    print(sheet.max_row, sheet.max_column)

    # 读取单元格数据
    cell_value = sheet['A1'].value

    # 遍历行
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=10)):
        print('************ {} ************'.format(index))
        for cell in row:
            if cell.value is None:
                break
            print(cell.value)
        print('')
    print('遍历行完成')

    # 遍历列
    for index, column in enumerate(sheet.iter_cols(min_col=1, max_col=10)):
        print('************ {} ************'.format(index))
        for cell in column:
            if cell.value is None:
                break
            print(cell.value)
        print('')
    print('遍历列完成')

if __name__ == '__main__':
    """程序入口
    """
    
    #os.system('chcp 936 & cls')
    print('******** starting ********')
    start_time = time.time()
    
    main()

    end_time = time.time()
    print('process spend {} s.\n'.format(round(end_time - start_time, 3)))
